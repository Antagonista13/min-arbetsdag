import json
from datetime import datetime
from pathlib import Path

import openpyxl

INPUT_FILE = "Uppdrag.xlsx"
OUTPUT_FILE = "resursteam.json"
SHEET_NAME = "UPPDRAG"


def is_blank(value):
    return value is None or str(value).strip() == ""


def serialize_date(value):
    if value is None:
        return None

    if hasattr(value, "date"):
        return value.date().isoformat()

    return str(value).strip()


def is_request(row):
    """
    Räknas som förfrågan om UPPDRAGSBESKRIVNING
    innehåller värdet 'Förfrågan'.
    """

    description = str(
        row.get("UPPDRAGSBESKRIVNING") or ""
    )

    description = (
        description
        .replace("\xa0", " ")
        .strip()
        .lower()
    )

    return description == "förfrågan""


def is_active(row):
    """
    Aktiva uppdrag.

    Förfrågningar räknas inte som aktiva uppdrag.
    """

    if row.get("Avslut") is not False:
        return False

    if is_request(row):
        return False

    return any(
        not is_blank(row.get(key))
        for key in (
            "SKOLA",
            "ELEV",
            "UPPDRAGSBESKRIVNING"
        )
    )


def needs_follow_up(row):
    """
    Ett aktivt uppdrag behöver följas upp om
    ansvarig eller placering saknas.
    """

    return is_active(row) and (
        is_blank(row.get("ANSVARIG"))
        or is_blank(row.get("Placering"))
    )


def load_assignments(path=INPUT_FILE):

    workbook = openpyxl.load_workbook(
        path,
        data_only=True
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Saknar bladet '{SHEET_NAME}' i {path}"
        )

    sheet = workbook[SHEET_NAME]

    headers = [
        cell.value
        for cell in sheet[3]
    ]

    if not headers or "Avslut" not in headers:
        raise ValueError(
            "Kunde inte hitta rubrikerna i rad 3."
        )

    assignments = []
    requests = []

    for values in sheet.iter_rows(
        min_row=4,
        values_only=True
    ):

        row = dict(zip(headers, values))
        if row.get("UPPDRAGSBESKRIVNING"):
            print(
                "UPPDRAGSBESKRIVNING:",
                repr(row.get("UPPDRAGSBESKRIVNING"))
    )

        # Hoppa över helt tomma rader
        if not any(
            not is_blank(value)
            for value in values
        ):
            continue

        item = {
            "school": row.get("SKOLA") or "",
            "student": row.get("ELEV") or "",
            "grade": row.get("Årsk."),
            "description":
                row.get("UPPDRAGSBESKRIVNING") or "",
            "start":
                serialize_date(row.get("START")),
            "end":
                serialize_date(row.get("SLUT")),
            "priority":
                row.get("PRIO"),
            "responsible":
                row.get("ANSVARIG") or "",
            "placement":
                row.get("Placering") or "",
            "notes":
                row.get("ÖVRIGT") or "",
        }

        # FÖRFRÅGAN
        if is_request(row):

            item["needsFollowUp"] = False

            requests.append(item)

            continue

        # AKTIVT UPPDRAG
        if not is_active(row):
            continue

        item["needsFollowUp"] = needs_follow_up(row)

        assignments.append(item)

    return assignments, requests


def fetch_resursteam(
    input_file=INPUT_FILE,
    output_file=OUTPUT_FILE
):

    assignments, requests = load_assignments(
        input_file
    )

    result = {
        "updatedAt": datetime.now()
        .astimezone()
        .isoformat(),

        "summary": {
            "activeAssignments":
                len(assignments),

            "totalAssignments":
                len(assignments),

            "requests":
                len(requests),

            "followUp":
                sum(
                    1
                    for item in assignments
                    if item["needsFollowUp"]
                ),
        },

        "assignments": assignments,

        "requests": requests,
    }

    Path(output_file).write_text(
        json.dumps(
            result,
            ensure_ascii=False,
            indent=2
        ),
        encoding="utf-8",
    )

    print(
        f"Resursteam: "
        f"{len(assignments)} aktiva uppdrag, "
        f"{len(requests)} förfrågningar, "
        f"{result['summary']['followUp']} att följa upp."
    )

    print(
        f"Skrev {output_file}"
    )


if __name__ == "__main__":
    fetch_resursteam()
